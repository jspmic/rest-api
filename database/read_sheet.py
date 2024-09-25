import openpyxl
from pathlib import Path


# Environment constants(if any change is made, modify these values
#                   not the classes, unless it's a major change in the sheet)
# -------------------------------------------------------------------------------
PATH_ABS = Path(__file__).parent
PATH = str(PATH_ABS.joinpath("./worksheet.xlsx"))

# Column numbers for each field
STOCK_CENTRAL = 1
LIVRAISON_RETOUR = 2
INPUT = 3
PROGRAM = 4
TYPE_TRANSPORT = 5
DISTRICTS = 6


# These classes are to be changed if any mandatory change is made.
# Before you change them, think of a way to create new classes
# that inherit from these

class Worksheet:
    """ Reads from the Excel file to retrieve mandatory values """

    def __init__(self, path: str = PATH) -> None:
        try:
            self.work_obj = openpyxl.load_workbook(path)
        except Exception:
            raise ValueError("Probably an invalid path")
        self.sheet_obj = self.work_obj.active
        if self.work_obj is not None and self.sheet_obj is not None:
            self.rows = self.sheet_obj.max_row
            self.cols = self.sheet_obj.max_column
        else:
            raise ValueError("No worksheet")

    def stock_central(self) -> list[str]:
        """ Returns the *Stock Central* possible values """

        stock_central: list[str] = []

        # The first value of the list is the name of the column
        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=STOCK_CENTRAL)
                if cell.value is not None:
                    stock_central.append(str(cell.value))

        return stock_central

    def livraison_retour(self) -> list[str]:
        """ Returns the Livraison_Retour field value """

        liv_ret: list[str] = []

        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=LIVRAISON_RETOUR)
                if cell.value is not None:
                    liv_ret.append(str(cell.value))

        return liv_ret

    def input(self) -> list[str]:
        """ Returns the *Input* possible values """

        inputs: list[str] = []

        # The first value of the list is the name of the column
        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=INPUT)
                if cell.value is not None:
                    inputs.append(str(cell.value))

        return inputs

    def program(self) -> list[str]:
        """ Returns the *Program* possible values """

        program: list[str] = []

        # The first value of the list is the name of the column
        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=PROGRAM)
                if cell.value is not None:
                    program.append(str(cell.value))

        return program

    def type_transport(self) -> list[str]:
        """ Returns the *Type_transport* possible values """

        transport: list[str] = []

        # The first value of the list is the name of the column
        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=TYPE_TRANSPORT)
                if cell.value is not None:
                    transport.append(str(cell.value))

        return transport

    def districts(self) -> list[str]:
        """ Returns the *Districts* possible values """

        districts: list[str] = []

        # The first value of the list is the name of the column
        if self.sheet_obj is not None:
            for row in range(1, (self.rows)+1):
                cell = self.sheet_obj.cell(row=row, column=DISTRICTS)
                if cell.value is not None:
                    districts.append(str(cell.value))

        return districts

    def colline(self, district: str = "") -> dict:
        """
        Returns all the location available to the district(if provided).
        If no district is given, it returns all the available locations
        """

        locations: dict = dict()
        index: int = 0

        # The location field must be located after the Clients field
        if self.sheet_obj is not None:
            for col in range(DISTRICTS+1, (self.cols)+1):
                for row in range(1, (self.rows)+1):
                    cell = self.sheet_obj.cell(row=row, column=col)
                    if cell.value is not None:
                        if index == 0:
                            col_name = str(cell.value)
                        else:
                            locations[col_name] = locations.get(
                                    col_name, []) + [cell.value]
                        index += 1
                index = 0

        if district:
            return locations.get(district, {})

        return locations

    def __str__(self) -> str:
        return f"Rows: {self.rows}\tCols: {self.cols}"


if __name__ == "__main__":
    sheet = Worksheet()
    districts = sheet.districts()
    print(f"{districts[-1]} = {sheet.colline(districts[-1])}")
