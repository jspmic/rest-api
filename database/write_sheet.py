from openpyxl import Workbook, load_workbook
from pathlib import Path
from enum import Enum
import os

PATH_ABS = Path(__file__).parent
PATH = str(PATH_ABS.joinpath("."))
os.chdir(PATH)


class TransfertColumns(Enum):
    ID = 1
    DATE = 2
    PLAQUE = 3
    LOGISTIC_OFFICIAL = 4
    NUMERO_MOUVEMENT = 5
    STOCK_CENTRAL_DEPART = 6


class LivraisonColumns(Enum):
    ID = 1
    DATE = 2
    PLAQUE = 3
    LOGISTIC_OFFICIAL = 4
    NUMERO_MOUVEMENT = 5
    STOCK_CENTRAL_DEPART = 6


class Writesheet:
    def __init__(self, f_name: str | None = None) -> None:
        self.workbook = load_workbook(f_name)
        self.sheet = self.workbook.active
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column

    def save(self, f_name: str = "") -> None:
        if f_name != "":
            name = f_name
        elif f_name == "" and self.f_name is not None:
            name = self.f_name
        else:
            raise ValueError("Please specify the filename in the arguments")

        self.workbook.save(filename=name)

    def insert_transfert(self, id: int, date: str,
                         plaque: str, logistic_official: str,
                         numero_mvt: int, stock_central_depart: str,
                         stock_central_suivant: dict,
                         stock_central_retour: str,
                         photo_mvt: str, type_transport: str,
                         motif: str
                         ):
        # for row in range(
        # self.sheet.cell(row=1
        pass


if __name__ == "__main__":
    w_s = Writesheet()
    w_s.save("dreams.xlsx")
    print(w_s.rows)
    print(w_s.cols)
    exit(0)
